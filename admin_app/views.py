import calendar
from calendar import month_name
from decimal import Decimal
from collections import defaultdict
from datetime import date
import json
import logging
from urllib.parse import urlencode
from django.db import transaction
from commerce.services.returns import process_refund_to_wallet
from django.shortcuts import render,redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import login,logout
from django.contrib.auth import authenticate
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.views.decorators.cache import never_cache,cache_control
from django.core.exceptions import ValidationError
from django.contrib import messages
from django.utils.timezone import now
from django.core.paginator import Paginator
from django.db.models import Q,Prefetch
from cloudinary.utils import cloudinary_url
from django.db.models.functions import TruncMonth, TruncYear,ExtractWeek
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Count,Sum,F
from django.utils import timezone
from datetime import datetime, timedelta
from openpyxl import Workbook
from django.template.loader import render_to_string
from weasyprint import HTML
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


from admin_app.services.sales_report import get_date_range, get_sold_items
from commerce.utils.coupons import calculate_item_coupon_share

from commerce.models import Orders,OrderItem,OrderReturn
from users.models import User,UserManager
from product.models import Category,Product,ProductVariant,ProductImage,ProductOffer,CategoryOffer,Coupon,CouponUsage
from product.forms import CategoryForm,ProductForm,ProductVariantForm,ProductOfferForm,CategoryOfferForm,CouponForm
from commerce.utils.orders import is_first_successful_order
from commerce.utils.referral import process_referral_after_first_order 
from commerce.services.returns import approve_return_service
from .services.order_status import update_order_payment_status,update_order_item_status

logger = logging.getLogger("admin_app")


def admin_login(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        user = authenticate(request,email=email,password=password)
        if user is not None:
            if user.is_admin:
                login(request,user)
                return redirect('admin_dashboard')
            else:
                messages.error(request,"You are not autherized to access admin panel.")
        else:
            messages.error(request,"Invalid email or password.")
    return render(request,'admin/admin_login.html')


@never_cache
@login_required(login_url='admin_login')
def admin_dashboard(request):
    if not request.user.is_admin:
        messages.error(request,"Unautherized access!")
        return redirect('admin_login')
    current_year = now().year
    years = range(current_year - 5, current_year + 1)
    months = list(month_name)[1:]

    return render(request, "admin/dashboard.html", {
        "years": years,
        "months": months,
        "current_month": now().month,
        "current_year": current_year,
    })

@login_required(login_url='admin_login')
def order_chart_data(request):
    filter_type = request.GET.get("filter", "monthly")
    year = request.GET.get("year")
    month = request.GET.get("month")

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    qs = Orders.objects.filter(
        payment_status__in=["paid", "partially_refunded"]
    )

    if start_date and end_date:
        qs = qs.filter(
            created_at__date__range=[start_date, end_date]
        )

        data = (
            qs.annotate(day=TruncMonth("created_at"))
              .values("day")
              .annotate(order_count=Count("id"))
              .order_by("day")
        )

        labels = [d["day"].strftime("%b %Y") for d in data]
        values = [d["order_count"] for d in data]

        return JsonResponse({
            "labels": labels,
            "values": values,
            "current_month": now().month
        })

    if not year:
        return JsonResponse({"labels": [], "values": []})

    year = int(year)

    if filter_type == "weekly":
        if not month:
            return JsonResponse({"labels": [], "values": []})

        month = int(month)

        orders = qs.filter(
            created_at__year=year,
            created_at__month=month
        )

        weekly_data = defaultdict(int)

        for order in orders:
            day = order.created_at.day
            week_of_month = (day - 1) // 7 + 1
            weekly_data[week_of_month] += 1

        labels = [f"Week {i}" for i in range(1, 6)]
        values = [weekly_data.get(i, 0) for i in range(1, 6)]

    elif filter_type == "monthly":
        data = (
            qs.filter(created_at__year=year)
              .annotate(month=TruncMonth("created_at"))
              .values("month")
              .annotate(order_count=Count("id"))
              .order_by("month")
        )

        monthly_data = {m: 0 for m in range(1, 13)}
        for entry in data:
            monthly_data[entry["month"].month] = entry["order_count"]

        labels = [calendar.month_abbr[m] for m in range(1, 13)]
        values = [monthly_data[m] for m in range(1, 13)]

    else:
        data = (
            qs.annotate(year=TruncYear("created_at"))
              .values("year")
              .annotate(order_count=Count("id"))
              .order_by("year")
        )

        labels = [entry["year"].year for entry in data]
        values = [entry["order_count"] for entry in data]

    return JsonResponse({
        "labels": labels,
        "values": values,
        "current_month": now().month
    })

def apply_order_time_filters(qs, request):
    filter_type = request.GET.get("filter")
    year = request.GET.get("year")
    month = request.GET.get("month")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if start_date and end_date:
        return qs.filter(
            order__created_at__date__range=[start_date, end_date]
        )

    if filter_type == "weekly" and year and month:
        return qs.filter(
            order__created_at__year=year,
            order__created_at__month=month
        )

    if filter_type == "monthly" and year:
        return qs.filter(order__created_at__year=year)

    if filter_type == "yearly" and year:
        return qs.filter(order__created_at__year=year)

    return qs

@login_required(login_url='admin_login')
def best_selling_categories_chart(request):
    qs = OrderItem.objects.filter(
        order__payment_status__in=["paid", "partially_refunded"],
        status="delivered"
    )

    qs = apply_order_time_filters(qs, request)

    data = (
        qs.values("product__product__category__name")
          .annotate(
              quantity_sold=Sum("quantity"),
              revenue=Sum(F("quantity") * F("price"))
          )
          .order_by("-quantity_sold")[:10]
    )

    return JsonResponse({
        "categories": [
            {
                "name": item["product__product__category__name"],
                "quantity": item["quantity_sold"],
                "revenue": float(item["revenue"] or 0)
            }
            for item in data
            if item["product__product__category__name"]
        ]
    })

    
@login_required(login_url='admin_login')
def best_selling_material_types_chart(request):
    qs = OrderItem.objects.filter(
        order__payment_status__in=["paid", "partially_refunded"],
        status="delivered"
    )
    qs = apply_order_time_filters(qs, request)

    data = (
        qs.values("product__material_type")
          .annotate(
              quantity_sold=Sum("quantity"),
              revenue=Sum(F("quantity") * F("price"))
          )
          .order_by("-quantity_sold")[:10]
    )

    return JsonResponse({
        "materials": [
            {
                "name": item["product__material_type"],
                "quantity": item["quantity_sold"],
                "revenue": float(item["revenue"] or 0)
            }
            for item in data
            if item["product__material_type"]
        ]
    })

@login_required(login_url='admin_login')
def best_selling_products_chart(request):
    qs = OrderItem.objects.filter(
        order__payment_status__in=["paid", "partially_refunded"],
        status="delivered"
    )

    qs = apply_order_time_filters(qs, request)

    data = (
        qs.values("product__product__name")
          .annotate(
              quantity_sold=Sum("quantity"),
              revenue=Sum(F("quantity") * F("price"))
          )
          .order_by("-quantity_sold")[:10]
    )

    return JsonResponse({
        "products": [
            {
                "name": item["product__product__name"],
                "quantity": item["quantity_sold"],
                "revenue": float(item["revenue"] or 0)
            }
            for item in data
            if item["product__product__name"]
        ]
    })



@never_cache
def admin_logout(request):
    logout(request)
    return redirect('admin_login')

@login_required(login_url='admin_login')
def customer_list(request):
    search_query = request.GET.get('q', '')
    filter_status = request.GET.get('filter', '')

    customers = User.objects.all().order_by('-created_at')

    if search_query:
        customers = customers.filter(Q(first_name__icontains=search_query)|Q(last_name__icontains=search_query))

    if filter_status == 'blocked':
        customers = customers.filter(is_blocked=True)
    elif filter_status == 'unblocked':
        customers = customers.filter(is_blocked=False)

    paginator = Paginator(customers, 12)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)

    return render(request, 'admin/customers.html', {
        'page_obj': page_obj,
        'customers':page_obj.object_list,
        'search_query': search_query,
        'filter_status': filter_status
    })

@login_required(login_url='admin_login')
def toggle_block_status(request, customer_id):
    if request.method=='POST':
        customer =get_object_or_404(User,id=customer_id)
        customer.is_blocked = not customer.is_blocked
        customer.save()
        return JsonResponse({
            'success':True,
            'is_blocked':customer.is_blocked,
            'customer_id':customer.id
        })
    return JsonResponse({'success':False})


@login_required(login_url='admin_login')
def admin_category_list(request):
    search_query=request.GET.get('q','')

    categories = Category.objects.all_with_deleted().order_by('is_deleted','-created_at')

    if search_query:
        categories = categories.filter(Q(name__icontains=search_query))

    paginator = Paginator(categories, 8)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    params = request.GET.copy()

    if "page" in params:
        params.pop("page")  
    querystring = params.urlencode()


    primary_images = {}
    for cat in page_obj:
        if cat.image:
            url , _= cloudinary_url(
                cat.image,
                width=100,
                height=100,
                crop='fill',
                gravity='auto',
                secure=True
            )
            primary_images[cat.id]=url

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'primary_images':primary_images,
        "querystring": querystring,

    }

    return render(request,'admin/category_list.html',context)



@login_required(login_url='admin_login')
def add_category(request):
    if request.method=='POST':
        print("POST:", request.POST)
        print("IMAGE RECEIVED:", request.POST.get("image"))

        form = CategoryForm(request.POST)
        print("FORM ERRORS:", form.errors)

        if form.is_valid():
            form.save()
            return redirect('admin_category_list')
    else:
        form=CategoryForm()
       
    return render(request,'admin/category_form.html',{'form':form})


@login_required(login_url='admin_login')
def edit_category(request,id):
    category =get_object_or_404(Category.objects.all_with_deleted(),id=id)
    if request.method=='POST':
        form = CategoryForm(request.POST or None, instance=category,show_deleted=True)
        if form.is_valid():
            form.save()
            return redirect('admin_category_list')
    else:
        form=CategoryForm(instance=category)

    if category.image:
        category_image_url, _ = cloudinary_url(
        category.image,
        width=300,
        height=300,
        crop="fill",
        secure=True
    )
    else:
        category_image_url = None
        
    context={
        'form':form,
        'category_image_url':category_image_url
    }
    return render(request,'admin/category_form.html',context)


@login_required(login_url='admin_login')
def delete_category(request, id):
    Category.objects.soft_delete(id)
    return redirect('admin_category_list')

@login_required(login_url='admin_login')
@require_POST
def restore_category(request, id):
    category = Category.objects.restore(id)
    if category:
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'message': 'Category not found or already active'})



@login_required(login_url='admin_login')
def admin_product_list(request):

    if request.GET.get('clear'):
        return redirect('admin_product_list')
    
   
    search_query = request.GET.get('q', '').strip()

    
    products = Product.objects.all_with_deleted().prefetch_related(
        Prefetch('images', queryset=ProductImage.objects.order_by('-is_primary', '-created_at'))
    ).order_by('is_deleted','-created_at')

    if search_query:
        products = products.filter(Q(name__icontains=search_query))


    paginator = Paginator(products, 9)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    params = request.GET.copy()
    if "page" in params:
        params.pop("page")
    clean_querystring = params.urlencode()



    primary_images = {}
    for product in page_obj:
        primary_image = product.images.filter(is_primary=True).first() or product.images.first()
        if primary_image and hasattr(primary_image.image, 'url'):
            primary_images[product.id] = primary_image.image.url
        else:
            primary_images[product.id] = None


    context = {
        'page_obj': page_obj,
        'products': page_obj.object_list,
        'search_query': search_query,
        'primary_images': primary_images, 
        "querystring": clean_querystring,

    }

    return render(request, 'admin/product_list.html', context)


@login_required(login_url='admin_login')
def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        image_urls = json.loads(request.POST.get('images', '[]'))


        if len(image_urls) < 3:
            messages.error(request, "Please upload at least 3 images before saving.")
            return render(request, 'admin/product_form.html', {'form': form})

        if form.is_valid():
            product = form.save()


            for url in image_urls:
                ProductImage.objects.create(product=product, image=url)

            messages.success(request, "Product added successfully.")
            request.session['product_saved'] = True
            return redirect('admin_variant_list', product_id=product.id)
        else:
            messages.error(request, "Please fix the errors below.")
    else:
        form = ProductForm()
        request.session.pop('product_saved', None)

    return render(request, 'admin/product_form.html', {'form': form})



@login_required(login_url='admin_login')
def edit_product(request, id):

    product = get_object_or_404(Product.objects.all_with_deleted(), id=id)

    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product, show_deleted=True)
        image_urls = json.loads(request.POST.get('images', '[]'))

        if len(image_urls) < 3:
            messages.error(request, "Please keep at least 3 product images.")
            return render(request, 'admin/product_form.html', {
                'form': form,
                'product': product,
                'existing_images': json.dumps(image_urls),
            })

        if form.is_valid():
            form.save()

            
            existing_images = [
                img.image.url
                for img in ProductImage.objects.filter(product=product)
                if hasattr(img.image, 'url')
            ]

            new_urls = set(image_urls)
            old_urls = set(existing_images)

            
            for removed_url in old_urls - new_urls:
                ProductImage.objects.filter(product=product, image__contains=removed_url).delete()

            
            for added_url in new_urls - old_urls:
                ProductImage.objects.create(product=product, image=added_url)

            messages.success(request, "Product updated successfully.")
            request.session['product_saved'] = True
            return redirect('admin_product_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = ProductForm(instance=product, show_deleted=True)
        request.session.pop('product_saved', None)

    existing_images = [
        img.image.url
        for img in ProductImage.objects.filter(product=product)
        if hasattr(img.image, 'url')
    ]

    context = {
        'form': form,
        'product': product,
        'existing_images': json.dumps(existing_images),
    }
    return render(request, 'admin/product_form.html', context)


@login_required(login_url='admin_login')
def delete_product(request, id):

    product = Product.objects.soft_delete(id)

    if not product:
        messages.warning(request, "⚠️ This product was already deleted or not found.")
    else:
        messages.success(request, f"🗑️ Product '{product.name}' deleted successfully.")

    return redirect('admin_product_list')



@login_required(login_url='admin_login')
@require_POST
def restore_product(request, id):
    product = Product.objects.restore(id)
    if product:
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'message': 'Product not found or already active'})


@login_required(login_url='admin_login')
def admin_variant_list(request, product_id):

    product = get_object_or_404(Product, id=product_id)
    search_query = request.GET.get('q', '').strip()
    filter_status = request.GET.get('filter', '')


    variants = ProductVariant.objects.all_with_deleted().filter(product=product)


    if search_query:
        variants = variants.filter(material_type__icontains=search_query) 

    if filter_status == 'active':
        variants = variants.filter(is_deleted=False)
    elif filter_status == 'deleted':
        variants = variants.filter(is_deleted=True)


    paginator = Paginator(variants.order_by('-created_at'), 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    params = request.GET.copy()
    if "page" in params:
        params.pop("page")
    clean_querystring = params.urlencode()


    return render(request, 'admin/variant_list.html', {
        'variants': page_obj,
        'page_obj': page_obj,
        'product': product,
        'search_query': search_query,
        'filter_status': filter_status,
        "querystring": clean_querystring,

    })

@login_required(login_url='admin_login')
def add_variant(request, product_id):

    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        form = ProductVariantForm(request.POST,instance=ProductVariant(product=product))
        if form.is_valid():
            variant = form.save(commit=False)
            variant.product = product  
            variant.save()
            messages.success(request, "Variant added successfully.")
            return redirect('admin_variant_list', product_id=product.id)
    else:
        form = ProductVariantForm()

    return render(request, 'admin/variant_form.html', {
        'form': form,
        'product': product
    })

@login_required(login_url='admin_login')
def edit_variant(request, id):
    variant = get_object_or_404(ProductVariant.objects.all_with_deleted(), id=id)
    product = variant.product  

    if request.method == 'POST':
        form = ProductVariantForm(request.POST, instance=variant,show_deleted=True)
        if form.is_valid():
            form.save()
            messages.success(request, "Variant updated successfully.")
            return redirect('admin_variant_list', product_id=product.id)
    else:
        form = ProductVariantForm(instance=variant,show_deleted=True)

    return render(request, 'admin/variant_form.html', {
        'form': form,
        'variant': variant,
        'product': product
    })

@login_required(login_url='admin_login')
def delete_variant(request, id):
    variant = ProductVariant.objects.soft_delete(id)
    if variant:
        messages.success(request, f"Variant '{variant.material_type}' deleted successfully.")
        return redirect('admin_variant_list', product_id=variant.product.id)
    messages.error(request, "Variant not found.")
    return redirect('admin_product_list')


@login_required(login_url='admin_login')
@require_POST
def restore_variant(request, id):
    variant = ProductVariant.objects.restore(id)
    if variant:
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'message': 'Variant not found or already active.'})


@login_required(login_url='admin_login')
def admin_order_list(request):
    search_query = request.GET.get("q", "")

    orders=Orders.objects.select_related("user","address").order_by("-created_at")

    if search_query:
        orders = orders.filter(order_id__icontains=search_query)

    paginator = Paginator(orders, 12)  
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "admin/order_list.html", {
        "page_obj": page_obj,
        "search_query": search_query,
    })

@cache_control(no_store=True, no_cache=True, must_revalidate=True)
@login_required(login_url='admin_login')
def admin_order_details(request, order_id):
    order = get_object_or_404(Orders, order_id=order_id)
    items = order.items.select_related("product", "product__product")

    if request.method == "POST":

        # payment status update
        if "payment_status" in request.POST:
            new_payment_status = request.POST["payment_status"]

            logger.info(
                f"[ADMIN] {request.user.email} updating payment status "
                f"Order={order.order_id} Status={new_payment_status}"
            )

            try:
                update_order_payment_status(order, new_payment_status)
                messages.success(request, "Payment status updated.")
            except ValidationError as e:
                logger.warning(
                    f"[PAYMENT BLOCKED] Order={order.order_id} Reason={str(e)}"
                )
                messages.error(request, e.messages[0])

            return redirect("order_details", order_id=order_id)

        # item status update
        item_id = request.POST.get("item_id")
        new_status = request.POST.get("status")

        item = get_object_or_404(OrderItem, id=item_id, order=order)

        logger.info(
            f"[ADMIN] {request.user.email} attempting item status update "
            f"Order={order.order_id} Item={item.id} "
            f"From={item.status} To={new_status}"
        )

        try:
            update_order_item_status(item, new_status)
            messages.success(request, "Order item status updated successfully.")

            logger.info(
                f"[ITEM UPDATED] Order={order.order_id} "
                f"Item={item.id} Status={new_status}"
            )

        except ValidationError as e:
            logger.warning(
                f"[ITEM UPDATE BLOCKED] Order={order.order_id} "
                f"Item={item.id} Reason={str(e)}"
            )
            messages.error(request,e.messages[0])

        return redirect("order_details", order_id=order_id)

    return render(
        request,
        "admin/order_details.html",
        {"order": order, "items": items}
    )
@require_POST
def approve_return(request, return_id):
    return_request = get_object_or_404(
        OrderReturn,
        id=return_id,
        approval_status="pending"
    )
    refund_amount = approve_return_service(return_request)

    messages.success(
        request,
        f"Return approved. ₹{refund_amount} refunded to wallet."
    )
    return redirect("admin_return_list")


@require_POST
def reject_return(request, return_id):
    return_request = get_object_or_404(
        OrderReturn,
        id=return_id,
        approval_status="pending"
    )

    admin_note = request.POST.get("admin_note", "").strip()

    return_request.approval_status = "rejected"
    return_request.admin_note = admin_note
    return_request.save(update_fields=["approval_status", "admin_note"])

    messages.success(request, "Return request rejected.")
    return redirect("admin_return_list")
  
@login_required(login_url='admin_login')
def admin_return_list(request):
    returns = (
        OrderReturn.objects
        .select_related("user", "item__order")
        .order_by("-created_at")
    )
    paginator = Paginator(returns, 10)  
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "admin/returns/return_list.html",
        {"returns": page_obj,
         "page_obj":page_obj}
    )

@login_required(login_url="admin_login")
def admin_offer_list(request):
    search_query = request.GET.get("q", "").strip()

    product_qs = ProductOffer.objects.all().order_by("-created_at")
    category_qs = CategoryOffer.objects.all().order_by("-created_at")

    if search_query:
        product_qs = product_qs.filter(
            Q(name__icontains=search_query) |
            Q(product__name__icontains=search_query)
        )
        category_qs = category_qs.filter(
            Q(name__icontains=search_query) |
            Q(category__name__icontains=search_query)
        )

    product_paginator = Paginator(product_qs, 8)
    category_paginator = Paginator(category_qs, 8)

    product_page = request.GET.get("product_page")
    category_page = request.GET.get("category_page")

    product_offers = product_paginator.get_page(product_page)
    category_offers = category_paginator.get_page(category_page)

    return render(request, "admin/offers/offer_list.html", {
        "product_offers": product_offers,
        "category_offers": category_offers,
        "search_query": search_query,
    })

@login_required(login_url="admin_login")
def admin_product_offer_create(request):
    form = ProductOfferForm(request.POST or None)
    if request.method == "POST":
        print("RAW POST start_date:", request.POST.get("start_date"))


    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Product offer created.")
        return redirect("admin_offer_list")

    return render(request, "admin/offers/product_offer_form.html", {
        "title": "Add Product Offer",
        "form": form,
    })
@login_required(login_url="admin_login")
def admin_category_offer_create(request):
    form = CategoryOfferForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Category offer created.")
        return redirect("admin_offer_list")

    return render(request, "admin/offers/category_offer_form.html", {
        "title": "Add Category Offer",
        "form": form,
    })


@login_required(login_url='admin_login')
def admin_product_offer_edit(request, pk):
    offer = get_object_or_404(ProductOffer, pk=pk)
    form = ProductOfferForm(request.POST or None, instance=offer)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Product offer updated.")
        return redirect("admin_offer_list")

    return render(request, "admin/offers/offer_form.html", {
        "title": "Edit Product Offer",
        "show_offer_type": False,   
        "offer_type": "product",
        "product_form": form,      
        "category_form": None,      
    })


@login_required(login_url='admin_login')
def admin_category_offer_edit(request, pk):
    offer = get_object_or_404(CategoryOffer, pk=pk)
    form = CategoryOfferForm(request.POST or None, instance=offer)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Category offer updated.")
        return redirect("admin_offer_list")

    return render(request, "admin/offers/offer_form.html", {
        "title": "Edit Category Offer",
        "show_offer_type": False,   
        "offer_type": "category",
        "product_form": None,
        "category_form": form,
    })

@login_required(login_url='admin_login')
@require_POST
def admin_offer_toggle(request, offer_type, pk):
    model = ProductOffer if offer_type == "product" else CategoryOffer
    offer = get_object_or_404(model, pk=pk)
    now = timezone.localtime(timezone.now())

    if not offer.is_active:
        conflict_qs = model.objects.filter(
            is_active=True,
            start_date__lte=now,
            end_date__gte=now
        ).exclude(pk=offer.pk)

        if offer_type == "product":
            conflict_qs = conflict_qs.filter(product=offer.product)
        else:
            conflict_qs = conflict_qs.filter(category=offer.category)

        if conflict_qs.exists():
            messages.error(
                request,
                "An active offer already exists for this "
                f"{'product' if offer_type == 'product' else 'category'}."
            )
            return redirect("admin_offer_list")

    offer.is_active = not offer.is_active
    offer.save(update_fields=["is_active"])

    status = "activated" if offer.is_active else "deactivated"
    messages.success(request, f"Offer {status} successfully.")
    return redirect("admin_offer_list")

@login_required(login_url="admin_login")
def delete_product_offer(request,pk):
    offer=get_object_or_404(ProductOffer,id=pk)
    if offer:
        offer.delete()
        messages.success(request,"Offer deletion successfull!")
    else:
        messages.error(request,"Currently unavailable!")
    return redirect("admin_offer_list")


@login_required(login_url="admin_login")
def delete_category_offer(request,pk):
    offer=get_object_or_404(CategoryOffer,id=pk)
    if offer:
        offer.delete()
        messages.success(request,"Offer deletion successfull!")
    else:
        messages.error(request,"Currently unavailable!")
    return redirect("admin_offer_list")

@login_required(login_url="admin_login")
def admin_coupon_list(request):
    search_query = request.GET.get("q", "")
    coupons = (
        Coupon.objects
        .all()
        .annotate(used_count=Count("usages"))
        .order_by("-created_at")
    )
    if search_query:
        coupons = coupons.filter(coupon_code__icontains=search_query)

    paginator = Paginator(coupons, 6)  
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)      
    return render(request, "admin/coupons/list.html", {
        "coupons": page_obj,
        "page_obj":page_obj
    })

@login_required(login_url="admin_login")
def admin_coupon_create(request):
    form = CouponForm(request.POST or None)

    if form.is_valid():
        form.save()
        messages.success(request, "Coupon created successfully")
        return redirect("admin_coupon_list")

    return render(request, "admin/coupons/form.html", {
        "form": form,
        "title": "Create Coupon"
    })

@login_required(login_url="admin_login")
def admin_coupon_edit(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk, is_deleted=False)
    form = CouponForm(request.POST or None, instance=coupon)

    if form.is_valid():
        form.save()
        messages.success(request, "Coupon updated successfully")
        return redirect("admin_coupon_list")

    return render(request, "admin/coupons/form.html", {
        "form": form,
        "title": "Edit Coupon"
    })
@login_required(login_url="admin_login")
def admin_coupon_toggle(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk)
    coupon.is_active = not coupon.is_active
    coupon.save(update_fields=["is_active"])

    messages.success(request, "Coupon status updated")
    return redirect("admin_coupon_list")

@login_required(login_url="admin_login")
def admin_coupon_delete(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk)
    coupon.is_deleted = True
    coupon.is_active = False
    coupon.save(update_fields=["is_deleted", "is_active"])

    messages.success(request, "Coupon deleted")
    return redirect("admin_coupon_list")


@login_required(login_url="admin_login")
def sales_report(request):

    range_type = request.GET.get("range")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if start_date and end_date:
        if start_date>end_date:
            messages.error(request,"Start date should be before end date")
            return redirect("sales_report")
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        range_type = None
    else:
        range_type = range_type or "daily"
        start, end = get_date_range(range_type, None, None)

    start, end = get_date_range(range_type, start_date, end_date)
    sold_items = get_sold_items(start, end)

    total_sales = Decimal("0.00")
    product_discount = Decimal("0.00")
    coupon_discount = Decimal("0.00")

    for item in sold_items:
        total_sales += item.price
        product_discount += (item.unit_price * item.quantity) - item.price
        coupon_discount += calculate_item_coupon_share(item.order, item)
    
    paginator = Paginator(sold_items,8)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    params = request.GET.copy()
    params.pop("page", None)
    querystring = urlencode(params)

    context = {
        "sold_items": page_obj,
        "page_obj":page_obj,
        "querystring": querystring,
        "order_count": sold_items.values("order_id").distinct().count(),
        "total_sales": total_sales,
        "product_discount": product_discount,
        "coupon_discount": coupon_discount,
        "overall_discount": product_discount + coupon_discount,
        "start_date": start,
        "end_date": end,
        "range_type": range_type,
    }

    return render(request, "admin/reports/sales_report.html", context)


@login_required(login_url="admin_login")
def sales_report_excel(request):

    range_type = request.GET.get("range")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    start, end = get_date_range(range_type, start_date, end_date)
    sold_items = get_sold_items(start, end)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    center_align = Alignment(horizontal="center")
    right_align = Alignment(horizontal="right")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    total_fill = PatternFill("solid", fgColor="FEE2E2")
    total_font = Font(bold=True)


    headers = [
        "Order ID",
        "Order Date",
        "Product",
        "Quantity",
        "Net Price (₹)",
        "Product Discount (₹)",
        "Coupon Discount (₹)",
    ]

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = 20

    ws.freeze_panes = "A2"

    total_sales = Decimal("0.00")
    product_discount = Decimal("0.00")
    coupon_discount = Decimal("0.00")

    row_num = 2

    for item in sold_items:
        prod_discount = (item.unit_price * item.quantity) - item.price
        coup_discount = calculate_item_coupon_share(item.order, item)

        ws.append([
            item.order.order_id,
            item.order.created_at.date(),
            str(item.product.product),
            item.quantity,
            float(item.price),
            float(prod_discount),
            float(coup_discount),
        ])

        ws.cell(row=row_num, column=2).number_format = "YYYY-MM-DD"
        ws.cell(row=row_num, column=4).alignment = center_align

        for col in (5, 6, 7):
            ws.cell(row=row_num, column=col).number_format = '₹#,##0.00'
            ws.cell(row=row_num, column=col).alignment = right_align

        for col in range(1, 8):
            ws.cell(row=row_num, column=col).border = thin_border

        total_sales += item.price
        product_discount += prod_discount
        coupon_discount += coup_discount

        row_num += 1


    row_num += 1

    totals = [
        ("TOTAL SALES", total_sales),
        ("PRODUCT DISCOUNT", product_discount),
        ("COUPON DISCOUNT", coupon_discount),
        ("OVERALL DISCOUNT", product_discount + coupon_discount),
    ]

    for label, value in totals:
        ws.append([label, "", "", "", float(value)])
        for col in range(1, 6):
            cell = ws.cell(row=row_num, column=col)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border

        ws.cell(row=row_num, column=5).number_format = '₹#,##0.00'
        ws.cell(row=row_num, column=5).alignment = right_align
        row_num += 1

   
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="sales_report.xlsx"'

    wb.save(response)
    return response

@login_required(login_url='admin_login')
def sales_report_pdf(request):
    range_type = request.GET.get("range", "daily")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    today = timezone.now().date()

    if range_type == "daily":
        start_date = end_date = today

    elif range_type == "weekly":
        start_date = today - timedelta(days=6)
        end_date = today

    elif range_type == "monthly":
        start_date = today.replace(day=1)
        end_date = today

    sold_items = OrderItem.objects.filter(
        status="delivered",
        order__created_at__date__range=(start_date, end_date)
    ).select_related("order", "product")

    total_sales = sum(i.price * i.quantity for i in sold_items)

    html = render_to_string(
        "admin/reports/sales_report_pdf.html",
        {
            "sold_items": sold_items,
            "start_date": start_date,
            "end_date": end_date,
            "total_sales": total_sales,
        }
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "inline; filename=sales_report.pdf"

    HTML(string=html).write_pdf(response)
    return response


FINAL_ITEM_STATUSES = {
    "cancelled",
    "delivered",
    "returned",
    "failed",
}

@transaction.atomic
def admin_cancel_order(request, order_id):
    order = get_object_or_404(Orders, order_id=order_id)
    if order.items.filter(status__in=["delivered","returned"]).exists():
        messages.error(
            request, 
            "This order cannot be cancelled because one or more items have already been completed. "
            "Please use the Return process instead."
        )
        return redirect("order_details", order_id=order.order_id)

    cancellable_items = order.items.exclude(status__in=FINAL_ITEM_STATUSES)
    if not cancellable_items.exists():
        messages.error(request,"Order cannot be cancelled. All items are already completed.")
        return redirect("order_details", order_id=order.order_id)

    if request.method == "POST":
        reason = request.POST.get("reason", "").strip()
        final_refund_amount = Decimal('0.00')

        if order.payment_status in ["paid","partially_refunded"]:

            total_paid = order.original_total
            already_refunded = order.refunded_amount
            final_refund_amount = total_paid - already_refunded
            if final_refund_amount > 0:
                process_refund_to_wallet(order, final_refund_amount, source="order_cancel")
            order.payment_status = "refunded"
        else:
            # For COD or Failed payments
            order.payment_status = "cancelled"        
        for item in cancellable_items.select_for_update():
            
            item.status = "cancelled"
            item.cancellation_reason = reason
            item.save(update_fields=['status', 'cancellation_reason'])

            item.product.stock += item.quantity
            item.product.save(update_fields=['stock'])

        order.total_price = Decimal("0.00")
        order.save(update_fields=['payment_status', 'total_price'])

        messages.success(request, "Order cancelled successfully.")
        return redirect("order_details", order_id=order.order_id)

    return redirect("order_details", order_id=order.order_id)